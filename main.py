import sys
import os
import openpyxl
import numpy
import random
from PyQt5.QtWidgets import QMainWindow, QApplication, QWidget, QAction, QTableWidget, QTableWidgetItem, QVBoxLayout, QHBoxLayout, QSizePolicy
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import pyqtSlot

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt


class MainWindow(QWidget):

    def __init__(self, *args, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        self.title = 'Hello world'
        self.left = 0
        self.top = 0
        self.width = 300
        self.height = 200
        self.expensepath = "/Users/sinithleng/Dropbox/Expense/monthly_expense_report.xlsx"
        self.monthList = []
        self.valueList = []
        self.initUI()

    def initUI(self):
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)
        self.layout = QHBoxLayout()
        self.createTable()
        self.m = PlotCanvas(self, width=6, height=4, dpi=100,
                            category=self.monthList, value=self.valueList)
        self.layout.addWidget(self.tableWidget)
        self.layout.addWidget(self.m)
        self.setLayout(self.layout)
        self.show()

    def createTable(self):
        self.tableWidget = QTableWidget()
        if not os.path.isfile(self.expensepath):
            print("invalid path")
        else:
            self.wb_obj = openpyxl.load_workbook(
                self.expensepath, data_only=True)
            self.sheet_obj = self.wb_obj.active
            self.maxrow = self.sheet_obj.max_row
            self.maxcol = self.sheet_obj.max_column
            self.tableWidget.setRowCount(self.maxrow)
            self.tableWidget.setColumnCount(self.maxcol)
            for row in range(1, self.maxrow+1):
                for col in range(1, self.maxcol+1):
                    self.cell_obj = self.sheet_obj.cell(row=row, column=col)
                    self.tableWidget.setItem(
                        row-1, col-1, QTableWidgetItem(str(self.cell_obj.value)))

            for j in range(2, self.maxcol+1):
                self.month = self.sheet_obj.cell(row=1, column=j)
                self.monthList.append(self.month.value)

            for k in range(2, self.maxcol+1):
                self.expense = self.sheet_obj.cell(row=28, column=k)
                self.valueList.append(self.expense.value)


class PlotCanvas(FigureCanvas):
    def __init__(self, parent=None, width=5, height=4, dpi=100, category=[], value=[]):
        fig = Figure(figsize=(width, height), dpi=dpi)
        self.axes = fig.add_subplot(111)

        FigureCanvas.__init__(self, fig)
        self.setParent(parent)

        FigureCanvas.setSizePolicy(self,
                                   QSizePolicy.Expanding,
                                   QSizePolicy.Expanding)
        FigureCanvas.updateGeometry(self)
        self.plot(category, value)

    def plot(self, category=[], value=[]):
        objects = category
        y_pos = numpy.arange(len(objects))
        performance = value

        self.axes.bar(y_pos, performance, align='center', alpha=0.5, color='r')
        self.axes.set_xticks(y_pos)
        self.axes.set_title('Monthly Expense')
        self.axes.set_ylabel('USD($)')
        self.axes.set_xticklabels(objects, rotation=45)
        self.draw()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MainWindow()
    sys.exit(app.exec_())
